"""把人工录入的资源 xlsx 导入 resources 表。

用法:
    .venv/bin/python -m scripts.import_resources <xlsx 路径> [sheet 名]

字段:subject / category / knowledge_point / title / type / url /
     estimated_minutes / summary / source / tags
- (category, knowledge_point) 通过 concepts.aliases 反查 concept_code
- (title, url) 视作唯一键,重复跳过
- title / type / summary / estimated_minutes 必填,缺失行集中报错

注:导入完成后请运行 `.venv/bin/python -m scripts.init_rag` 重建向量索引。
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

from app.database import SessionLocal, init_db
from app.models import Concept, Resource


def _norm(value) -> str:
    return str(value).strip() if value is not None else ""


def _to_int(value, default: int | None = None) -> int | None:
    if value in (None, ""):
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _parse_tags(value) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return [str(t).strip() for t in value if str(t).strip()]
    return [t.strip() for t in str(value).split(",") if t.strip()]


def _build_alias_map(db) -> dict[tuple[str, str], str]:
    out: dict[tuple[str, str], str] = {}
    for c in db.query(Concept).all():
        for a in c.aliases or []:
            key = (_norm(a.get("category")), _norm(a.get("knowledge_point")))
            if all(key):
                out[key] = c.code
    return out


def _pick_sheet(wb: openpyxl.Workbook, sheet_name: str | None):
    if sheet_name and sheet_name in wb.sheetnames:
        return wb[sheet_name]
    for name in wb.sheetnames:
        if "录入" in name:
            return wb[name]
    return wb[wb.sheetnames[0]]


def import_xlsx(xlsx_path: Path, sheet_name: str | None = None) -> dict:
    init_db()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = _pick_sheet(wb, sheet_name)

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return {"success": 0, "skipped": 0, "failed": [], "note": "sheet 行数 < 3"}

    headers = [_norm(h) for h in rows[0]]
    header_idx = {h: i for i, h in enumerate(headers) if h}

    required = {"category", "knowledge_point", "title", "type", "summary", "estimated_minutes"}
    missing = required - set(header_idx)
    if missing:
        raise ValueError(f"xlsx 缺少必需列: {sorted(missing)},请用最新模板录入")

    success = 0
    skipped = 0
    failed: list[dict] = []

    with SessionLocal() as db:
        alias_map = _build_alias_map(db)

        for row_num, row in enumerate(rows[2:], start=3):
            if not any(row):
                continue
            data = {h: row[i] for h, i in header_idx.items() if i < len(row)}

            category = _norm(data.get("category"))
            kp = _norm(data.get("knowledge_point"))
            title = _norm(data.get("title"))
            rtype = _norm(data.get("type"))
            summary = _norm(data.get("summary"))
            est = _to_int(data.get("estimated_minutes"))

            if not category or not kp:
                failed.append({"row": row_num, "reason": "category/knowledge_point 为空"})
                continue
            if not title:
                failed.append({"row": row_num, "reason": "title 为空"})
                continue
            if not rtype:
                failed.append({"row": row_num, "reason": "type 为空"})
                continue
            if not summary:
                failed.append({"row": row_num, "reason": "summary 为空(RAG 检索质量靠它)"})
                continue
            if est is None or est <= 0:
                failed.append({"row": row_num, "reason": "estimated_minutes 必须为正整数"})
                continue

            concept_code = alias_map.get((category, kp))
            if not concept_code:
                failed.append({
                    "row": row_num,
                    "reason": f"未找到匹配的 concept_code: ({category} - {kp}),"
                              f"请对照 docs/概念清单.md 修正",
                })
                continue

            url = _norm(data.get("url")) or None

            existing = (
                db.query(Resource)
                .filter(Resource.title == title, Resource.url == url)
                .first()
            )
            if existing:
                skipped += 1
                continue

            db.add(
                Resource(
                    concept_code=concept_code,
                    title=title,
                    type=rtype,
                    url=url,
                    estimated_minutes=est,
                    summary=summary,
                )
            )
            success += 1

        db.commit()

    return {"success": success, "skipped": skipped, "failed": failed}


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("用法: python -m scripts.import_resources <xlsx 路径> [sheet 名]",
              file=sys.stderr)
        return 1
    xlsx_path = Path(argv[1]).expanduser()
    if not xlsx_path.exists():
        print(f"[ERR] 文件不存在: {xlsx_path}", file=sys.stderr)
        return 1
    sheet_name = argv[2] if len(argv) > 2 else None

    result = import_xlsx(xlsx_path, sheet_name)
    print(f"[OK] 成功 {result['success']} 条, 跳过(重复) {result['skipped']} 条, "
          f"失败 {len(result['failed'])} 条")
    if result["failed"]:
        print("--- 失败明细 ---")
        for f in result["failed"]:
            print(f"  行 {f['row']}: {f['reason']}")
    if result["success"] > 0:
        print("提示:导入成功后请运行 `.venv/bin/python -m scripts.init_rag` 重建向量索引")
    return 0 if not result["failed"] else 2


if __name__ == "__main__":
    sys.exit(main(sys.argv))
