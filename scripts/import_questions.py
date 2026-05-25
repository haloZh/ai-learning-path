"""把人工录入的题库 xlsx 导入 questions 表。

用法:
    .venv/bin/python -m scripts.import_questions <xlsx 路径> [sheet 名]

行为:
- 默认读 "录入表(同学填这里)" sheet,可通过参数覆盖
- 第 1 行为字段名(英文),第 2 行为中文说明,从第 3 行开始为题目数据
- (category, knowledge_point) -> concept_code 通过 concepts.aliases 反查
- 同 stem+source+year 视作重复,跳过(脚本可重复跑)
- 失败行集中报告,不影响成功行入库
- LaTeX 原文照存,不做任何转义
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

from app.database import SessionLocal, init_db
from app.models import Concept, Question

# JSON 标准只接受 \" \\ \/ \b \f \n \r \t \uXXXX 这些 escape;
# 题库里 LaTeX 命令 \dfrac \sqrt \pi 等会让 json.loads 失败,
# 这里把所有"非法 escape"的单反斜杠转成双反斜杠后再试一次。
_INVALID_BACKSLASH = re.compile(r'\\(?!["\\/bfnrtu])')

DEFAULT_SHEET = "录入表(同学填这里)"
ALT_SHEET = "录入表(同学填这里)"


def _norm(value) -> str:
    return str(value).strip() if value is not None else ""


def _to_int(value, default: int | None = None) -> int | None:
    if value in (None, ""):
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _parse_choices(value) -> dict | list:
    """options 字段:可能是 dict / JSON 字符串 / 空;兼容 LaTeX 反斜杠。"""
    if value in (None, ""):
        return {}
    if isinstance(value, (dict, list)):
        return value
    text = str(value).strip()
    if not text:
        return {}
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        # LaTeX 命令(\dfrac/\sqrt 等)在 JSON 里非法,转义后重试
        fixed = _INVALID_BACKSLASH.sub(r"\\\\", text)
        return json.loads(fixed)


def _parse_tags(value) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return [str(t).strip() for t in value if str(t).strip()]
    return [t.strip() for t in str(value).split(",") if t.strip()]


def _build_alias_map(db) -> dict[tuple[str, str], str]:
    out: dict[tuple[str, str], str] = {}
    for c in db.query(Concept).all():
        for a in c.aliases or []:
            key = (_norm(a.get("category")), _norm(a.get("knowledge_point")))
            if all(key):
                out[key] = c.code
    return out


# 每个 category 的兜底 concept_code:精确/substring 都未命中时使用,
# 保证 100% 入库;命中此层的题应被人工 review。
_CATEGORY_DEFAULT: dict[str, str] = {
    # 数据分析 / 几何
    "排列组合": "数据分析-组合",
    "数据描述": "数据分析-数据描述",
    "数列": "代数-数列-等差",
    "平面几何": "几何-平面几何-圆",
    "解析几何": "几何-解析几何-直线方程",
    "立体几何": "几何-空间几何体",
    "概率": "数据分析-古典概率",
    "容斥应用题": "数据分析-组合",
    # 算术 - 基础
    "绝对值模块": "算术-绝对值",
    "实数模块": "算术-绝对值",  # 多见"非负数之和为零"
    "平均值模块": "算术-平均数",
    "平均值统计模块": "算术-平均数",
    "比例应用题": "算术-比例",
    # 算术 - 应用题
    "行程应用题": "算术-应用题-行程",
    "工程应用题": "算术-应用题-工程",
    "浓度应用题": "算术-应用题-浓度",
    "利润应用题": "算术-应用题-利润",
    "年龄应用题": "代数-一元一次方程",  # 年龄差不变 → 等量关系建模
    # 代数 - 整式 / 分式
    "整式运算": "代数-整式运算",
    "整式乘法": "代数-整式运算",
    "整式求值": "代数-整式运算",
    "整式配方": "代数-整式运算",
    "因式分解": "代数-整式运算",
    "分式基础": "代数-分式运算",
    "分式化简": "代数-分式运算",
    "分式运算": "代数-分式运算",
    "分式恒等变形": "代数-分式运算",
    "分式方程": "代数-分式运算",
    # 代数 - 方程
    "一元一次方程": "代数-一元一次方程",
    "二元一次方程组": "代数-一元一次方程",
    "一元二次方程": "代数-一元二次方程",
    "含参一元二次方程": "代数-一元二次方程",
    "韦达定理": "代数-一元二次方程-韦达定理",
    # 代数 - 不等式
    "分式不等式": "代数-不等式",
    "一元一次不等式": "代数-不等式",
    "一元二次不等式": "代数-不等式",
    "高次不等式": "代数-不等式",
    "均值不等式不等式": "代数-不等式",
}


def _resolve_concept_code(
    category: str, kp: str, alias_map: dict[tuple[str, str], str]
) -> tuple[str | None, str]:
    """三层反查:精确 → substring(同 category) → category 兜底。

    返回 (concept_code, 命中策略),策略用于诊断/日志。
    """
    if (category, kp) in alias_map:
        return alias_map[(category, kp)], "exact"

    candidates = [
        (a_kp, code)
        for (a_cat, a_kp), code in alias_map.items()
        if a_cat == category
    ]
    candidates.sort(key=lambda x: -len(x[0]))
    for a_kp, code in candidates:
        if a_kp and a_kp in kp:
            return code, "substring"

    if category in _CATEGORY_DEFAULT:
        return _CATEGORY_DEFAULT[category], "category-default"
    return None, "miss"


def _pick_sheet(wb: openpyxl.Workbook, sheet_name: str | None):
    if sheet_name and sheet_name in wb.sheetnames:
        return wb[sheet_name]
    # 默认行为: 优先匹配名称里有 "录入" 的 sheet
    for name in wb.sheetnames:
        if "录入" in name:
            return wb[name]
    return wb[wb.sheetnames[0]]


def import_xlsx(xlsx_path: Path, sheet_name: str | None = None) -> dict:
    init_db()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = _pick_sheet(wb, sheet_name)

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return {"success": 0, "skipped": 0, "failed": [], "note": "sheet 行数 < 3,无可导入数据"}

    headers = [_norm(h) for h in rows[0]]
    header_idx = {h: i for i, h in enumerate(headers) if h}

    required = {"category", "knowledge_point", "content", "correct_answer"}
    missing = required - set(header_idx)
    if missing:
        raise ValueError(f"xlsx 缺少必需列: {sorted(missing)},请用最新模板录入")

    success = 0
    skipped = 0
    failed: list[dict] = []
    by_strategy: dict[str, int] = {"exact": 0, "substring": 0, "category-default": 0}
    fallback_rows: list[dict] = []  # 命中 substring 或 category-default 的行,供 review

    # 第 2 行可能是字段中文说明(原模板风格),也可能直接是数据(同学清理过模板)。
    # 用 difficulty 是否可解析为整数判断:中文说明会写"难度(1-5)",真题写 1~5。
    data_start = 2  # 默认跳过第 1 行 header 与第 2 行说明
    if len(rows) >= 2:
        row2_diff_idx = header_idx.get("difficulty")
        if row2_diff_idx is not None and row2_diff_idx < len(rows[1]):
            try:
                int(float(_norm(rows[1][row2_diff_idx])))
                data_start = 1  # 第 2 行 difficulty 是数字,认为已经是数据
            except (TypeError, ValueError):
                pass

    with SessionLocal() as db:
        alias_map = _build_alias_map(db)

        for row_num, row in enumerate(rows[data_start:], start=data_start + 1):
            if not any(row):
                continue

            data = {h: row[i] for h, i in header_idx.items() if i < len(row)}
            category = _norm(data.get("category"))
            knowledge_point = _norm(data.get("knowledge_point"))
            stem = _norm(data.get("content"))

            if not category or not knowledge_point:
                failed.append({"row": row_num, "reason": "category 或 knowledge_point 为空"})
                continue
            if not stem:
                failed.append({"row": row_num, "reason": "content 为空"})
                continue

            concept_code, strategy = _resolve_concept_code(
                category, knowledge_point, alias_map
            )
            if not concept_code:
                failed.append({
                    "row": row_num,
                    "reason": f"未找到匹配的 concept_code: ({category} - {knowledge_point}),"
                              f"请对照 docs/概念清单.md 修正",
                })
                continue

            by_strategy[strategy] = by_strategy.get(strategy, 0) + 1
            if strategy != "exact":
                fallback_rows.append({
                    "row": row_num,
                    "category": category,
                    "knowledge_point": knowledge_point,
                    "concept_code": concept_code,
                    "strategy": strategy,
                })

            try:
                choices = _parse_choices(data.get("options"))
            except json.JSONDecodeError as e:
                opt_preview = str(data.get("options") or "")[:60]
                failed.append({
                    "row": row_num,
                    "reason": f"options JSON 解析失败({e.msg});原文片段:{opt_preview!r}",
                })
                continue

            source = _norm(data.get("source")) or None
            year = _to_int(data.get("year"))

            existing = (
                db.query(Question)
                .filter(
                    Question.stem == stem,
                    Question.source == source,
                    Question.year == year,
                )
                .first()
            )
            if existing:
                skipped += 1
                continue

            db.add(
                Question(
                    concept_code=concept_code,
                    question_type=_norm(data.get("question_type")) or "single_choice",
                    stem=stem,
                    choices=choices,
                    answer=_norm(data.get("correct_answer")),
                    difficulty=_to_int(data.get("difficulty"), 1) or 1,
                    score=_to_int(data.get("score"), 3) or 3,
                    explanation=_norm(data.get("explanation")) or None,
                    source=source,
                    year=year,
                    tags=_parse_tags(data.get("tags")),
                )
            )
            success += 1

        db.commit()

    return {
        "success": success,
        "skipped": skipped,
        "failed": failed,
        "by_strategy": by_strategy,
        "fallback_rows": fallback_rows,
    }


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("用法: python -m scripts.import_questions <xlsx 路径> [sheet 名]", file=sys.stderr)
        return 1

    xlsx_path = Path(argv[1]).expanduser()
    if not xlsx_path.exists():
        print(f"[ERR] 文件不存在: {xlsx_path}", file=sys.stderr)
        return 1

    sheet_name = argv[2] if len(argv) > 2 else None

    result = import_xlsx(xlsx_path, sheet_name)
    print(f"[OK] 成功 {result['success']} 条, 跳过(重复) {result['skipped']} 条, "
          f"失败 {len(result['failed'])} 条")
    bs = result.get("by_strategy", {})
    if any(bs.values()):
        print(
            f"     命中策略: 精确 {bs.get('exact', 0)} | "
            f"substring {bs.get('substring', 0)} | "
            f"category 兜底 {bs.get('category-default', 0)}(建议 review)"
        )
    fb = result.get("fallback_rows", [])
    if fb:
        # 仅打印 category-default 的兜底,substring 通常足够准
        defaults_only = [r for r in fb if r["strategy"] == "category-default"]
        if defaults_only:
            print(f"--- category 兜底入库行(共 {len(defaults_only)},建议人工核对) ---")
            for r in defaults_only[:30]:
                print(f"  行 {r['row']}: ({r['category']} - {r['knowledge_point']}) → "
                      f"{r['concept_code']}")
            if len(defaults_only) > 30:
                print(f"  ...及另外 {len(defaults_only) - 30} 行(略)")
    if result["failed"]:
        print("--- 失败明细 ---")
        for f in result["failed"]:
            print(f"  行 {f['row']}: {f['reason']}")
    return 0 if not result["failed"] else 2


if __name__ == "__main__":
    sys.exit(main(sys.argv))
